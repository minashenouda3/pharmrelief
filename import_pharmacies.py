#!/usr/bin/env python3
"""
Load Ontario pharmacy data into pharmacy_directory.

Usage:
    export SUPABASE_DB_URL='postgresql://...'
    python3 import_pharmacies.py pharmacies.csv --dry-run
    python3 import_pharmacies.py pharmacies.csv

Handles CSV and Excel. Column names are matched loosely, so the usual
variations ("Pharmacy Name", "name", "OPERATING NAME") all resolve without
editing the file first.

Upserts on ocp_accreditation, so re-running with updated data refreshes
existing rows rather than duplicating them. Rows already claimed by a real
account keep their claim.
"""

import os
import sys
import csv
import re
import argparse

try:
    import psycopg2
    from psycopg2.extras import execute_batch
except ImportError:
    sys.exit("pip install psycopg2-binary")

# Loose header matching: lowercased, non-alphanumerics stripped.
FIELD_ALIASES = {
    'ocp_accreditation': ['ocp', 'accreditation', 'accreditationnumber', 'acc',
                          'ocpnumber', 'licencenumber', 'licensenumber', 'pharmacyid'],
    'name':              ['name', 'pharmacyname', 'operatingname', 'businessname',
                          'tradename', 'storename'],
    'legal_name':        ['legalname', 'corporatename', 'registeredname'],
    'address_line':      ['address', 'addressline', 'address1', 'streetaddress', 'street'],
    'city':              ['city', 'town', 'municipality'],
    'province':          ['province', 'prov', 'state'],
    'postal_code':       ['postalcode', 'postal', 'zip', 'zipcode'],
    'phone':             ['phone', 'telephone', 'phonenumber', 'tel'],
    'fax':               ['fax', 'faxnumber'],
    'email':             ['email', 'emailaddress'],
    'website':           ['website', 'url', 'web'],
    'banner':            ['banner', 'chain', 'brand', 'group'],
    'pharmacy_type':     ['type', 'pharmacytype', 'category', 'classification'],
    'designated_manager':['designatedmanager', 'manager', 'dm', 'pharmacistincharge',
                          'pic', 'designatedmanagername'],
}

def norm(h):
    return re.sub(r'[^a-z0-9]', '', (h or '').lower())

def build_map(headers):
    """Map source columns onto directory fields."""
    normed = {norm(h): h for h in headers}
    mapping, used = {}, set()
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normed and normed[alias] not in used:
                mapping[field] = normed[alias]
                used.add(normed[alias])
                break
    return mapping

def clean(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None

def clean_postal(v):
    v = clean(v)
    if not v:
        return None
    v = re.sub(r'\s+', '', v).upper()
    return f"{v[:3]} {v[3:]}" if len(v) == 6 else v

def clean_phone(v):
    v = clean(v)
    if not v:
        return None
    d = re.sub(r'\D', '', v)
    if len(d) == 11 and d.startswith('1'):
        d = d[1:]
    return f"({d[:3]}) {d[3:6]}-{d[6:]}" if len(d) == 10 else v

def read_rows(path):
    if path.lower().endswith(('.xlsx', '.xls')):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("pip install openpyxl")
        ws = load_workbook(path, read_only=True, data_only=True).active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else '' for h in next(rows)]
        return headers, [dict(zip(headers, r)) for r in rows]
    # csv, sniffing the delimiter
    with open(path, newline='', encoding='utf-8-sig') as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        return reader.fieldnames or [], list(reader)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('file')
    ap.add_argument('--dry-run', action='store_true',
                    help='parse and report without writing')
    ap.add_argument('--source', default='OCP public register')
    ap.add_argument('--limit', type=int, help='only process the first N rows')
    args = ap.parse_args()

    headers, rows = read_rows(args.file)
    if args.limit:
        rows = rows[:args.limit]
    print(f"Read {len(rows)} rows, {len(headers)} columns")

    mapping = build_map(headers)
    print("\nColumn mapping:")
    for field in FIELD_ALIASES:
        src = mapping.get(field)
        print(f"  {field:<20} <- {src if src else '(not found)'}")

    unmapped = [h for h in headers if h not in mapping.values()]
    if unmapped:
        print(f"\nUnmapped source columns (ignored): {unmapped}")

    if 'name' not in mapping:
        sys.exit("\nERROR: no pharmacy name column found. Rename it to 'Name' and re-run.")
    if 'ocp_accreditation' not in mapping:
        print("\nWARNING: no OCP accreditation column. Rows cannot be deduplicated on "
              "re-import, and running this twice will create duplicates.")

    records, skipped = [], 0
    seen = set()
    for r in rows:
        get = lambda f: clean(r.get(mapping[f])) if f in mapping else None
        name = get('name')
        if not name:
            skipped += 1
            continue
        acc = get('ocp_accreditation')
        if acc:
            if acc in seen:
                skipped += 1
                continue
            seen.add(acc)
        records.append((
            acc, name, get('legal_name'), get('address_line'), get('city'),
            get('province') or 'ON', clean_postal(get('postal_code')),
            clean_phone(get('phone')), clean_phone(get('fax')),
            get('email'), get('website'), get('banner'),
            get('pharmacy_type'), get('designated_manager'), args.source
        ))

    print(f"\nPrepared {len(records)} records ({skipped} skipped: no name or duplicate)")
    if records:
        print("\nFirst 3:")
        for rec in records[:3]:
            print(f"  {rec[1]} | {rec[4]} | acc={rec[0]}")

    if args.dry_run:
        print("\nDry run — nothing written.")
        return

    dsn = os.environ.get('SUPABASE_DB_URL')
    if not dsn:
        sys.exit("Set SUPABASE_DB_URL first.")

    sql = """
    insert into pharmacy_directory
      (ocp_accreditation, name, legal_name, address_line, city, province,
       postal_code, phone, fax, email, website, banner, pharmacy_type,
       designated_manager, source)
    values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    on conflict (ocp_accreditation) do update set
      name = excluded.name,
      legal_name = excluded.legal_name,
      address_line = excluded.address_line,
      city = excluded.city,
      postal_code = excluded.postal_code,
      phone = excluded.phone,
      fax = excluded.fax,
      email = excluded.email,
      website = excluded.website,
      banner = excluded.banner,
      pharmacy_type = excluded.pharmacy_type,
      designated_manager = excluded.designated_manager,
      source = excluded.source
    -- a claimed listing keeps its claim; the pharmacy owns its own record
    where pharmacy_directory.claimed_by is null;
    """

    conn = psycopg2.connect(dsn)
    try:
        with conn, conn.cursor() as cur:
            execute_batch(cur, sql, records, page_size=500)
            cur.execute("select count(*), count(claimed_by) from pharmacy_directory")
            total, claimed = cur.fetchone()
        print(f"\nDone. Directory now holds {total} pharmacies ({claimed} claimed).")
    finally:
        conn.close()

if __name__ == '__main__':
    main()
